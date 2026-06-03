#!/usr/bin/env python3
"""
Generate the per-year ColDP metadata files for the colac (CoL Annual Checklist) source
from the editor's compilation spreadsheets in resources/colac/citations/:
  - ACs_metadata_summary_v1.xlsx — per-year title/description/scope/editor list/etc.
  - ACs_editors00-19_v1.xlsx     — per-year, per-editor full ColDP Agent records (orcid +
    organisation/department/city/state/country); populated for 2000 & 2002-2005 only.

Writes src/main/resources/colac/metadata/<year>.yaml — one explicit, hand-editable file per year.
Two placeholders remain for the generator to fill at runtime: {issued} and the trailing {sources}.

Re-run only if the spreadsheet or the author map changes; otherwise edit the YAML files directly.
"""
import openpyxl, os, sys

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "src/main/resources/colac/citations/ACs_metadata_summary_v1.xlsx")
# Per-year, per-editor full ColDP Agent records (orcid + affiliation). Columns map 1:1 to the
# ColDP Agent type. Only the 2000 & 2002-2005 sheets carry data; 2006-2019 are headers only,
# so editors in those years fall back to the AUTHORS map + affiliation() below.
EDITORS_XLSX = os.path.join(ROOT, "src/main/resources/colac/citations/ACs_editors00-19_v1.xlsx")
OUT  = os.path.join(ROOT, "src/main/resources/colac/metadata")

# Country name -> ISO 3166-1 alpha-2, for the editor spreadsheet's free-text Country column.
COUNTRY_ISO = {
    "philippines": "PH", "uk": "GB", "united kingdom": "GB", "germany": "DE",
    "australia": "AU", "usa": "US", "united states": "US", "france": "FR",
    "netherlands": "NL", "the netherlands": "NL",
}

# Fix obvious typos in the hand-compiled editor spreadsheet's organisation/city text.
TEXT_FIXES = {"Reasearch": "Research", "Liebniz": "Leibniz", "Los-Baños": "Los Baños"}

def fix_text(s):
    for bad, good in TEXT_FIXES.items():
        s = s.replace(bad, good)
    return s

def iso(country):
    c = country.strip()
    return COUNTRY_ISO.get(c.casefold(), c)

def load_editor_details():
    """(year, family.casefold()) -> dict of ColDP Agent fields from ACs_editors00-19_v1.xlsx.
    Columns: 0 Family 1 Given 2 email 3 ORCID 4 Organisation 5 RORID 6 Department 7 City
    8 State 9 Country 10 Url 11 Note."""
    wb = openpyxl.load_workbook(EDITORS_XLSX, data_only=True)
    out = {}
    for sheet in wb.sheetnames:
        try: year = int(sheet)
        except ValueError: continue
        for r in list(wb[sheet].iter_rows(values_only=True))[1:]:   # skip header
            cell = lambda i: ("" if i >= len(r) or r[i] is None else str(r[i]).strip())
            family = cell(0)
            if not family: continue
            out[(year, family.casefold())] = {
                "orcid":        cell(3),
                "organisation": fix_text(cell(4)),
                "department":   fix_text(cell(6)),
                "city":         fix_text(cell(7)),
                "state":        cell(8),
                "country":      iso(cell(9)) if cell(9) else "",
            }
    return out

# Expanded author names; ORCIDs only where verified (this repo's prior map + dataset/3).
# Given names + affiliations for Kimani, Hernandez and Paglinawan confirmed by the editor (2026);
# affiliations are emitted per editor (see affiliation()).
AUTHORS = {
    "Abucay L.":            ("Luisa", "Abucay", None),
    "Appeltans W.":         ("Ward", "Appeltans", None),
    "Baillargeon G.":       ("Guy", "Baillargeon", None),
    "Bailly N.":            ("Nicolas", "Bailly", "0000-0003-4994-0653"),
    "Bisby F.":             ("Frank A.", "Bisby", None),
    "Bisby F.A.":           ("Frank A.", "Bisby", None),
    "Bourgoin T.":          ("Thierry", "Bourgoin", None),
    "Brewer P.W.":          ("Peter W.", "Brewer", None),
    "Cachuela-Palacio M.":  ("Monalisa", "Cachuela-Palacio", None),
    "Culham A.":            ("Alastair", "Culham", None),
    "De Wever A.":          ("Aaike", "De Wever", None),
    "DeWalt R.E.":          ("R. Edward", "DeWalt", "0000-0001-9985-9250"),
    "Decock W.":            ("Wim", "Decock", "0000-0002-2168-9471"),
    "Didžiulis V.":         ("Viktoras", "Didžiulis", None),
    "Flann C.":             ("Christina", "Flann", None),
    "Froese R.":            ("Rainer", "Froese", "0000-0001-9745-636X"),
    "Hernandez F.":         ("Francisco", "Hernandez", None),
    "Kimani S.":            ("Susana", "Kimani", None),
    "Kimani S.W.":          ("Susana", "Kimani", None),
    "Kirk P.":              ("Paul M.", "Kirk", "0000-0002-0658-7338"),
    "Kirk P.M.":            ("Paul M.", "Kirk", "0000-0002-0658-7338"),
    "Kunze T.":             ("Thomas", "Kunze", None),
    "Nicolson D.":          ("David", "Nicolson", "0000-0002-7987-0679"),
    "Nieukerken E. van":    ("Erik J.", "van Nieukerken", None),
    "Orrell T.":            ("Thomas M.", "Orrell", "0000-0003-1038-3028"),
    "Orrell T.M.":          ("Thomas M.", "Orrell", "0000-0003-1038-3028"),
    "Ouvrard D.":           ("David", "Ouvrard", "0000-0003-2931-6116"),
    "Ower G.":              ("Geoff", "Ower", "0000-0002-9770-2345"),
    "Paglinawan L.":        ("Luvie", "Paglinawan", None),
    "Paglinawan L.E.":      ("Luvie", "Paglinawan", None),
    "Penev L.":             ("Lyubomir", "Penev", None),
    "Roskov Y.":            ("Yury R.", "Roskov", "0000-0003-2137-2690"),
    "Roskov Y.R.":          ("Yury R.", "Roskov", "0000-0003-2137-2690"),
    "Ruggiero M.":          ("Michael A.", "Ruggiero", None),
    "Ruggiero M.A.":        ("Michael A.", "Ruggiero", None),
    "Soulier-Perkins A.":   ("Adeline", "Soulier-Perkins", "0000-0002-9537-8537"),
    "Wilson K.":            ("Karen L.", "Wilson", "0000-0001-7419-8222"),
    "Wilson K.L.":          ("Karen L.", "Wilson", "0000-0001-7419-8222"),
    "Zarucchi J.":          ("James L.", "Zarucchi", None),
    "van Hertum J.":        ("Jorrit", "van Hertum", None),
}

# Editor affiliations confirmed by the editor (2026) for the three formerly-initialled editors.
# Returns (organisation, city, country) or None. Paglinawan moved institution, so it is year-keyed.
def affiliation(family, year):
    if family == "Kimani":
        return ("The University of Reading", "Reading", "GB")
    if family == "Hernandez":
        return ("Vlaams Instituut Voor De Zee", "Ostend", "BE")
    if family == "Paglinawan":
        if year <= 2010:
            return ("WorldFish Center", "Los Baños", "PH")
        return ("The FishBase Information and Research Group, Inc. (FIN)", "Los Baños", "PH")
    return None

# publisher city/country per era (ISO country codes) from the spreadsheet
def publisher(year):
    if year <= 2004: city, country = "Los Baños", "PH"
    elif year <= 2013: city, country = "Reading", "GB"
    else: city, country = "Leiden", "NL"
    org = "Species 2000 Catalogue of Life" if year == 2000 else "Species 2000 & ITIS Catalogue of Life"
    return org, city, country

KEYWORDS = ["COL", "Catalogue of Life", "species list", "community"]

def sq(s):  # single-quoted YAML scalar
    return "'" + s.replace("'", "''") + "'"

def block(s, indent):
    pad = " " * indent
    out = []
    for line in s.split("\n"):
        out.append(pad + line if line.strip() else "")
    return "\n".join(out)

def parse_editors(s):
    people = []
    for tok in s.split(","):
        tok = tok.strip()
        if not tok: continue
        if tok in AUTHORS:
            given, family, orcid = AUTHORS[tok]
        else:                                    # defensive fallback: "Surname I.I."
            parts = tok.rsplit(" ", 1)
            family, given, orcid = (parts[0], parts[1], None) if len(parts) == 2 else (tok, None, None)
        people.append((given, family, orcid))
    return people

def conversion_desc(year):
    base = ("Converted to ColDP from the original Species 2000/ITIS Annual Checklist MySQL "
            "database with the export scripts in this repository.")
    if year <= 2004:
        base += (" These early checklists were originally distributed as MS Access databases, "
                 "which were first imported into MySQL before conversion.")
    return base

def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    details = load_editor_details()
    ws = wb["Sheet1"]; rows = list(ws.iter_rows(values_only=True))
    years = [int(c) for c in rows[0] if c is not None]
    F = {str(r[0]): [("" if c is None else str(c)) for c in r[1:1+len(years)]] for r in rows[1:] if r[0]}
    os.makedirs(OUT, exist_ok=True)
    for i, year in enumerate(years):
        g = lambda f: F.get(f, [""]*len(years))[i]
        org, city, country = publisher(year)
        L = []
        L.append("# yaml-language-server: $schema=https://raw.githubusercontent.com/CatalogueOfLife/coldp/master/metadata.json")
        L.append("---")
        L.append(f"alias: {g('Alias (in CLB)')}")
        L.append(f"title: {sq(g('Title'))}")
        L.append("description: |-")
        L.append(block(g('Description'), 2))
        L.append("issued: {issued}")
        L.append(f"version: Annual Checklist {year}")
        L.append(f"issn: {g('ISSN')}")
        L.append("keyword:")
        for k in KEYWORDS: L.append(f"  - {k}")
        L.append(f"url: http://www.catalogueoflife.org/annual-checklist/{year}/")
        L.append(f"taxonomicScope: {g('Taxonomic scope') or 'Biota'}")
        L.append(f"geographicScope: {g('Geographic scope') or 'Global'}")
        L.append(f"temporalScope: {g('Temporal scope') or 'Extant taxa'}")
        L.append(f"confidence: {g('Checklist Confidence') or 4}")
        if g('Completeness (% of est. extant spp)'):
            L.append(f"completeness: {g('Completeness (% of est. extant spp)')}")
        L.append("license: CC BY")
        L.append("contact:")
        L.append("  organisation: COL Secretariat")
        L.append("  city: Amsterdam")
        L.append("  country: NL")
        L.append("  email: support@catalogueoflife.org")
        L.append("publisher:")
        L.append(f"  organisation: {org}")
        L.append(f"  city: {sq(city)}")
        L.append(f"  country: {country}")
        L.append("editor:")
        for given, family, orcid in parse_editors(g('Editor')):
            det = details.get((year, family.casefold()))
            L.append(f"  - family: {family}")
            if given: L.append(f"    given: {given}")
            eff_orcid = (det and det["orcid"]) or orcid
            if eff_orcid: L.append(f"    orcid: {eff_orcid}")
            if det:                                  # full per-year affiliation from the editor sheet
                if det["organisation"]: L.append(f"    organisation: {sq(det['organisation'])}")
                if det["department"]:   L.append(f"    department: {sq(det['department'])}")
                if det["city"]:         L.append(f"    city: {sq(det['city'])}")
                if det["state"]:        L.append(f"    state: {sq(det['state'])}")
                if det["country"]:      L.append(f"    country: {det['country']}")
            else:                                    # fall back to the 3 confirmed affiliations
                aff = affiliation(family, year)
                if aff:
                    org, city, country = aff
                    L.append(f"    organisation: {sq(org)}")
                    L.append(f"    city: {sq(city)}")
                    L.append(f"    country: {country}")
        L.append("contributor:")
        L.append("  - family: Döring")
        L.append("    given: Markus")
        L.append("    orcid: 0000-0001-7757-1889")
        L.append("    country: DE")
        L.append("    city: Berlin")
        L.append("    email: mdoering@gbif.org")
        L.append("    note: Developer of the ColDP generator code")
        L.append("conversion:")
        L.append("  url: https://github.com/CatalogueOfLife/coldp-generator")
        L.append("  description: >-")
        L.append(block(conversion_desc(year), 4))
        L.append("notes: |-")
        L.append(block(
            f"Generated from the historical CoL Annual Checklist MySQL database col{year}ac.\n"
            "IDs: t<id> accepted taxa, a<id> accepted names without classification, "
            "s<id> synonyms (and bare names), r<id> references, d<id> source databases (GSDs).\n"
            "Each NameUsage.sourceID points to its contributing Global Species Database.", 2))
        L.append("{sources}")
        path = os.path.join(OUT, f"{year}.yaml")
        with open(path, "w", encoding="utf-8") as fh:
            fh.write("\n".join(L) + "\n")
        print(f"wrote {path}  ({len(F['Editor'][i].split(','))} editors)")

if __name__ == "__main__":
    main()
